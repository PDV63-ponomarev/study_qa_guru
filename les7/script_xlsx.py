# библиотеки для xlsx
# pip install openpyxl

from openpyxl import load_workbook

workbook = load_workbook(r'tmp\excel.xlsx')

# работа с листом
sheet = workbook.active

# прочитать 3 строчку 2 колонки
print(sheet.cell(row = 3, column = 2).value)


